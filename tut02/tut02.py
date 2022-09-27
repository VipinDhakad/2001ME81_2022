# tut2 solution
import pandas as pd
import openpyxl
inp = openpyxl.load_workbook('input_octant_transition_identify.xlsx')
sheet_input = inp.active

output = openpyxl.Workbook()
out = output.active
mr = sheet_input.max_row
mc = sheet_input.max_column
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# reading cell value from source excel file
		data = sheet_input.cell(row = i, column = j)

		# writing the read value to destination excel file
		out.cell(row = i, column = j).value = data.value

output.save("output_octant_transition_identify.xlsx")

def octant_transition_count(mod=5000):
    data=pd.read_excel('output_octant_transition_identify.xlsx')
    u_avg=data['U'].mean()
    v_avg=data['V'].mean()
    w_avg=data['W'].mean()
    data.at[0,'U_avg']=u_avg
    data.at[0,'V_avg']=v_avg
    data.at[0,'W_avg']=w_avg
    i=0
    cnt_p1=0
    cnt_n1=0
    cnt_p2=0
    cnt_n2=0
    cnt_p3=0
    cnt_n3=0
    cnt_p4=0
    cnt_n4=0

    for ele in data['U']:
        x=data.at[i,"U'=U-Uavg"]=data.at[i,'U']-u_avg
        y=data.at[i,"V'=V-Vavg"]=data.at[i,'V']-v_avg
        z=data.at[i,"W'=W-Wavg"]=data.at[i,'W']-w_avg
        if x>0:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=1
                    cnt_p1=cnt_p1+1
                else:
                    data.at[i,'Octant']=-1
                    cnt_n1=cnt_n1+1
            else:
                if z>0:
                    data.at[i,'Octant']=4
                    cnt_p4=cnt_p4+1
                else:
                    data.at[i,'Octant']=-4
                    cnt_n4=cnt_n4+1
        else:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=2
                    cnt_p2=cnt_p2+1
                else:
                    data.at[i,'Octant']=-2
                    cnt_n2=cnt_n2+1
            else:
                if z>0:
                    data.at[i,'Octant']=3
                    cnt_p3=cnt_p3+1
                else:
                    data.at[i,'Octant']=-3
                    cnt_n3=cnt_n3+1
        i=i+1
    data.at[0,'Octant ID']='Overall Count'
    data.at[0,'1']=cnt_p1
    data.at[0,'-1']=cnt_n1
    data.at[0,'2']=cnt_p2
    data.at[0,'-2']=cnt_n2
    data.at[0,'3']=cnt_p3
    data.at[0,'-3']=cnt_n3
    data.at[0,'4']=cnt_p4
    data.at[0,'-4']=cnt_n4

    mod=5000
    i=0
    prev=0
    iter=1
    cnt_p1,cnt_n1,cnt_p2,cnt_n2,cnt_p3,cnt_n3,cnt_p4,cnt_n4=0,0,0,0,0,0,0,0
    data.at[1,'Octant ID']=mod
    # print(len(data))
    while i<len(data):
        flag=False
        for j in range(prev,mod*iter):
            if i<len(data):
                if data.at[i,'Octant']==1:
                    cnt_p1=cnt_p1+1
                if data.at[i,'Octant']==-1:
                    cnt_n1=cnt_n1+1
                if data.at[i,'Octant']==2:
                    cnt_p2=cnt_p2+1
                if data.at[i,'Octant']==-2:
                    cnt_n2=cnt_n2+1
                if data.at[i,'Octant']==3:
                    cnt_p3=cnt_p3+1
                if data.at[i,'Octant']==-3:
                    cnt_n3=cnt_n3+1
                if data.at[i,'Octant']==4:
                    cnt_p4=cnt_p4+1
                if data.at[i,'Octant']==-4:
                    cnt_n4=cnt_n4+1
                i=i+1
            else:
                flag=True
        data.at[iter+1,'Octant ID']=str(prev)+"-"+str(mod*iter-1)
        if flag:
            data.at[iter+1,'Octant ID']=str(prev)+"-"+str(len(data))
        data.at[iter+1,'1']=cnt_p1
        data.at[iter+1,'-1']=cnt_n1
        data.at[iter+1,'2']=cnt_p2
        data.at[iter+1,'-2']=cnt_n2
        data.at[iter+1,'3']=cnt_p3
        data.at[iter+1,'-3']=cnt_n3
        data.at[iter+1,'4']=cnt_p4
        data.at[iter+1,'-4']=cnt_n4
        cnt_p1,cnt_n1,cnt_p2,cnt_n2,cnt_p3,cnt_n3,cnt_p4,cnt_n4=0,0,0,0,0,0,0,0
        prev=mod*iter
        iter=iter+1
    data.at[13,'Octant ID']='Overall Transition Count'
    points=[1,-1,2,-2,3,-3,4,-4]
    for i in range(14,22):
        data.at[15,f'{points[i-14]}']=points[i-14]
    for i in range(16,24):
        data.at[i,'Octant ID']=points[i-16]
    print(data.iloc[0,1])
    for i in range(16,24):
        for j in range(12,20):
            data.iloc[i,j]=0
    for i in range(0,mr-2):
        row_val=0
        col_val=0
        if data.iloc[i,10]==1:
            row_val=16
        elif data.iloc[i,10]==-1:
            row_val=17
        elif data.iloc[i,10]==2:
            row_val=18
        elif data.iloc[i,10]==-2:
            row_val=19
        elif data.iloc[i,10]==3:
            row_val=20
        elif data.iloc[i,10]==-3:
            row_val=21
        elif data.iloc[i,10]==4:
            row_val=22
        elif data.iloc[i,10]==-4:
            row_val=23
        if data.iloc[i+1,10]==1:
            col_val=1
        elif data.iloc[i+1,10]==-1:
            col_val=-1
        elif data.iloc[i+1,10]==2:
            col_val=2
        elif data.iloc[i+1,10]==-2:
            col_val=-2
        elif data.iloc[i+1,10]==3:
            col_val=3
        elif data.iloc[i+1,10]==-3:
            col_val=-3
        elif data.iloc[i+1,10]==4:
            col_val=4
        elif data.iloc[i+1,10]==-4:
            col_val=-4
        # if pd.isnull(data.at[row_val, data.iloc[i+1,10]]):
        #     data.at[row_val, data.iloc[i+1,10]]=0
        data.at[row_val,f'{col_val}']+=1

    data.to_excel('output_octant_transition_identify.xlsx',index=False)
from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

mod=5000
octant_transition_count(mod)



