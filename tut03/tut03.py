#Help https://youtu.be/H37f_x4wAC0
# tut3 solution
import pandas as pd
import openpyxl
try:
    inp = openpyxl.load_workbook('input_octant_longest_subsequence.xlsx')
except:
    print("File not found!")
sheet_input = inp.active
##opening a new workbook to save as a output
output = openpyxl.Workbook()
out = output.active

##getting the size of the dataframe in terms of max rows and max columns
mr = sheet_input.max_row
mc = sheet_input.max_column

##copying the data of the input file into the newly created empty output file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# reading cell value from source input file
		data = sheet_input.cell(row = i, column = j)

		# writing the read value to destination output file
		out.cell(row = i, column = j).value = data.value

output.save("output_octant_longest_subsequence.xlsx")
def write_data():
    try:
        data=pd.read_excel('output_octant_longest_subsequence.xlsx')
    except:
        print("File not found!")
    ##finding the mean of the column "U", "V", and "W" with the help of mean functions
    u_avg=data['U'].mean()
    v_avg=data['V'].mean()
    w_avg=data['W'].mean()
    ##writing the data to the first cells in the output file
    data.at[0,'U_avg']=u_avg
    data.at[0,'V_avg']=v_avg
    data.at[0,'W_avg']=w_avg
    i=0
    ##calculating the values of deviation from the mean of every column
    for ele in data['U']:
        x=data.at[i,"U'=U-Uavg"]=data.at[i,'U']-u_avg
        y=data.at[i,"V'=V-Vavg"]=data.at[i,'V']-v_avg
        z=data.at[i,"W'=W-Wavg"]=data.at[i,'W']-w_avg
        ##tagging the octant begins here
        if x>0:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=1
                else:
                    data.at[i,'Octant']=-1
            else:
                if z>0:
                    data.at[i,'Octant']=4
                else:
                    data.at[i,'Octant']=-4
        else:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=2
                else:
                    data.at[i,'Octant']=-2
            else:
                if z>0:
                    data.at[i,'Octant']=3
                else:
                    data.at[i,'Octant']=-3
        i=i+1
    print("Data writing complete")
    ##saving the output file generated
    data.to_excel('output_octant_longest_subsequence.xlsx',index=False)
    ##longest subsequence count function definition
def octant_longest_subsequence_count():
    data=pd.read_excel('output_octant_longest_subsequence.xlsx')
    octants=[1,-1,2,-2,3,-3,4,-4]
    curr_max,overall_max=0,0
    data.at[0,""]=None
    for i in range(0,8):
        data.at[i,'Count']=octants[i]
    ##finding the size of longest subsequence and its frequency
    for octant in octants:
        freq=1
        for ele in data['Octant']:
            if ele==octant:
                curr_max+=1
                if curr_max==overall_max: freq+=1
                else: freq=1
                overall_max=max(overall_max,curr_max)
            else:
                curr_max=0
        row_val=0
        if octant>0:
            row_val=2*octant-2
        else:
            row_val=-2*octant-1
        ##updating the values accordingly to the excel sheet
        data.at[row_val,'Longest Subsquence Length']=overall_max
        data.at[row_val,'Frequency']=freq
        curr_max,overall_max=0,0
    print("Longest Subsequence length finding complete")
    data.to_excel('output_octant_longest_subsequence.xlsx',index=False)



from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

##calling the function to get our work done
try:
    write_data()
except:
    print("Error in copying data from input to newly generated output file")
try:
    octant_longest_subsequence_count()
except:
    print("Error in finding longest subsequence length")