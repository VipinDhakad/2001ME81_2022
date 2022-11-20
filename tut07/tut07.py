import pandas as pd
import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
from platform import python_version
from datetime import datetime
start_time = datetime.now()


octant_name_id_mapping = {'+1': "Internal outward interaction",
                          "-1": "External outward interaction",
                          "+2": "External Ejection",
                          "-2": "Internal Ejection",
                          "+3": "External inward interaction",
                          "-3": "Internal inward interaction",
                          "+4": "Internal sweep",
                          "-4": "External sweep"}
octant_key = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
yellow = PatternFill(start_color='00FFFF00',
                     end_color='00FFFF00',
                     fill_type='solid')
ifile_path = "E:\\GitHub\\2001ME81_2022\\tut07\\input\\"
ofile_path = "E:\\GitHub\\2001ME81_2022\\tut07\\output\\"
os.chdir(ifile_path)


def borderSetFunc(ws, cell_range):
    border_thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=border_thin, left=border_thin,
                                 right=border_thin, bottom=border_thin)


def checkFloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def generate_copy(input_workbook, mod):         #opening input files
   
    ifile = openpyxl.load_workbook(ifile_path + input_workbook)
    sheet_input = ifile.active
    sheet_output = openpyxl.Workbook()
    output = sheet_output.active
    mr = sheet_input.max_row
    mc = sheet_input.max_column
    for row in range(1, mr + 1):
        for col in range(1, mc + 1):   #reading input
            val = str(sheet_input.cell(row=row, column=col).value)  #writing to output
            if checkFloat(val) is True:
                output.cell(row=row, column=col).value = round(float(val), 3)
            else:
                output.cell(row=row, column=col).value = val
    output_sheet_name = input_workbook[:-5] + \
        ' _octant_analysis_mod_{}.xlsx'.format(mod)
    sheet_output.save(ofile_path + output_sheet_name)
    return ofile_path+output_sheet_name


def octant_tagging(sheet):
    mr = sheet.max_row
    sheet['E1'] = 'U Avg'
    sheet['F1'] = 'V Avg'
    sheet['G1'] = 'W Avg'
    sheet['H1'] = "U' = U - U Avg"
    sheet['I1'] = "V' = V - V Avg"
    sheet['J1'] = "W' = W - W Avg"
    sheet['K1'] = 'Octant'
    u_sum = 0
    v_sum = 0
    w_sum = 0
    for row in range(2, mr + 1):
        u_sum += sheet.cell(row=row, column=2).value
        v_sum += sheet.cell(row=row, column=3).value
        w_sum += sheet.cell(row=row, column=4).value
    u_mean = round(u_sum / (mr - 1), 3)
    v_mean = round(v_sum / (mr - 1), 3)
    w_mean = round(w_sum / (mr - 1), 3)
    sheet['E2'] = u_mean
    sheet['F2'] = v_mean
    sheet['G2'] = w_mean
    for row in range(2, mr + 1):
        sheet.cell(row=row, column=8).value = sheet.cell(
            row=row, column=2).value - u_mean
        sheet.cell(row=row, column=9).value = sheet.cell(
            row=row, column=3).value - v_mean
        sheet.cell(row=row, column=10).value = sheet.cell(
            row=row, column=4).value - w_mean
    p1,p2,p3,p4,n1,n2,n3,n4 = 0, 0, 0, 0, 0, 0, 0, 0
    for row in range(2, mr + 1):
        coord_x = sheet.cell(row=row, column=8).value
        coord_y = sheet.cell(row=row, column=9).value
        coord_z = sheet.cell(row=row, column=10).value
        if coord_x > 0:
            if coord_y > 0:
                if coord_z > 0:
                    sheet.cell(row=row, column=11).value = '+1'
                    p1 = p1 + 1
                else:
                    sheet.cell(row=row, column=11).value = '-1'
                    n1 = n1 + 1
            else:
                if coord_z > 0:
                    sheet.cell(row=row, column=11).value = '+4'
                    p4 = p4 + 1
                else:
                    sheet.cell(row=row, column=11).value = '-4'
                    n4 = n4 + 1
        else:
            if coord_y > 0:
                if coord_z > 0:
                    sheet.cell(row=row, column=11).value = '+2'
                    p2 = p2 + 1
                else:
                    sheet.cell(row=row, column=11).value = '-2'
                    n2 = n2 + 1
            else:
                if coord_z > 0:
                    sheet.cell(row=row, column=11).value = '+3'
                    p3 = p3 + 1
                else:
                    sheet.cell(row=row, column=11).value = '-3'
                    n3 = n3 + 1
        row += 1


def rank_and_count_func(mod, sheet):
    sheet['N1'] = 'Overall Octant Count'
    sheet['N3'] = 'Octant ID'
    sheet['N4'] = 'Overall Count'
    sheet['M4'] = 'Mod {}'.format(mod)
    count_freq_mapping = {'+1': 0, '-1': 0,      #storing the count and ranks
             '+2': 0, '-2': 0,
             '+3': 0, '-3': 0,
             '+4': 0, '-4': 0}
    mod_count_freq_mapping = {'+1': 0, '-1': 0,
                 '+2': 0, '-2': 0,
                 '+3': 0, '-3': 0,
                 '+4': 0, '-4': 0}
    ranking_count = dict()
    rank_octants = dict()
    counts = list()
    mr = sheet.max_row - 1
    for row in range(mr):
        count_freq_mapping[sheet.cell(row=2 + row, column=11).value] += 1   #total counts
    col = 0
    for row in octant_key:
        sheet.cell(row=3, column=15 + col).value = row
        sheet.cell(row=4, column=15 + col).value = count_freq_mapping[row]  #writing overall counts
        counts.append(count_freq_mapping[row])
        col = col + 1
    counts.sort(reverse=True)
    col = 1
    for row in counts:
        ranking_count[row] = col       #stroing count and rank in dict
        col += 1
    for row in octant_key:
        rank_octants[row] = ranking_count[count_freq_mapping[row]]
    col = 0
    for row in octant_key:
        sheet.cell(row=3, column=23 + col).value = 'Rank Octant {}'.format(row) #writing individual ranks
        sheet.cell(row=4, column=23 + col).value = rank_octants[row]
        col += 1

    iter = int(mr / mod) + 1
    p = int(0)
    q = int(mod - 1)
    for row in range(iter):
        ranking_count = dict()
        rank_octants = dict()
        counts = list()
        if(q > mr):
            q = mr - 1
            sheet.cell(row=5 + row, column=14).value = '{} - {}'.format(p, q + 1)
        else:
            sheet.cell(row=5 + row, column=14).value = '{} - {}'.format(p, q)
        for col in range(p, q + 1):
            mod_count_freq_mapping[sheet.cell(row=2 + col, column=11).value] += 1
        z = 0
        for ind in octant_key:
            sheet.cell(row=5 + row, column=15 + z).value = mod_count_freq_mapping[ind]
            counts.append(mod_count_freq_mapping[ind])
            z = z + 1
        counts.sort(reverse=True)
        col = 1
        for count in counts:
            ranking_count[count] = col
            col += 1
        for key in octant_key:
            rank_octants[key] = ranking_count[mod_count_freq_mapping[key]]   #wrinting corresponding ranks
        col = 0
        for t in octant_key:
            sheet.cell(row=5 + row, column=23 + col).value = rank_octants[t]
            col += 1
        p = p + int(mod)
        q = q + int(mod)
        mod_count_freq_mapping = {'+1': 0, '-1': 0,
                     '+2': 0, '-2': 0,
                     '+3': 0, '-3': 0,
                     '+4': 0, '-4': 0}

    sheet['AE3'] = 'Rank1 Octant ID'
    sheet['AF3'] = 'Rank1 Octant Name'
    col = 0
    for row in octant_key:
        if sheet.cell(row=4, column=23 + col).value == 1:
            sheet.cell(row=4, column=23 + col).fill = yellow
            sheet.cell(row=4, column=31).value = row
            sheet.cell(row=4, column=32).value = octant_name_id_mapping[row]
        col += 1

    for row in range(iter):
        ind = 0
        for col in octant_key:
            if sheet.cell(row=5 + row, column=23 + ind).value == 1:
                sheet.cell(row=5 + row, column=23 + ind).fill = yellow
                sheet.cell(row=5 + row, column=31).value = col
                sheet.cell(
                    row=5 + row, column=32).value = octant_name_id_mapping[col]
            ind += 1
    sheet['AC{}'.format(6 + iter)] = 'Octant ID'
    sheet['AD{}'.format(6 + iter)] = 'Octant Name'
    sheet['AE{}'.format(6 + iter)] = 'Count of Rank 1 in Mod Values'
    rank_count = {'+1': 0, '-1': 0,
                  '+2': 0, '-2': 0,
                  '+3': 0, '-3': 0,
                  '+4': 0, '-4': 0}
    for row in range(iter):
        oc = sheet.cell(row=5 + row, column=31).value
        rank_count[oc] += 1
    col = 0
    for row in octant_key:
        sheet.cell(row=7 + iter + col, column=29).value = row
        sheet.cell(row=7 + iter + col,
                   column=30).value = octant_name_id_mapping[row]
        sheet.cell(row=7 + iter + col, column=31).value = rank_count[row]
        col += 1
    borderSetFunc(sheet, 'N3:AF{}'.format(4 + iter))
    borderSetFunc(sheet, 'AC{}:AE{}'.format(6 + iter, 14 + iter))


def octant_transition_count(sheet):
    sheet['AI1'] = 'Overall Transition Count'
    sheet['AH4'] = 'From'
    sheet['AI3'] = 'Octant #'
    sheet['AJ2'] = 'To'
    mr = sheet.max_row - 1
    iter = int(mr / mod) + 1
    for row in range(8):
        sheet.cell(row=3, column=36 + row).value = octant_key[row]
        sheet.cell(row=4 + row, column=35).value = octant_key[row]
        for col in range(8):
            sheet.cell(row=4 + row, column=36 + col).value = 0
    for row in range(2, mr):
        index = 0
        count = 0
        val1 = sheet.cell(row=row, column=11).value
        val2 = sheet.cell(row=row + 1, column=11).value
        if val1 == '+1':
            index = 0
        elif val1 == '-1':
            index = 1
        elif val1 == '+2':
            index = 2
        elif val1 == '-2':
            index = 3
        elif val1 == '+3':
            index = 4
        elif val1 == '-3':
            index = 5
        elif val1 == '+4':
            index = 6
        elif val1 == '-4':
            index = 7
        if val2 == '+1':
            count = 0
        elif val2 == '-1':
            count = 1
        elif val2 == '+2':
            count = 2
        elif val2 == '-2':
            count = 3
        elif val2 == '+3':
            count = 4
        elif val2 == '-3':
            count = 5
        elif val2 == '+4':
            count = 6
        elif val2 == '-4':
            count = 7
        sheet.cell(row=4 + index, column=36 + count).value += 1
    p = 0
    q = mod - 1
    for row in range(iter):
        sheet.cell(row=15 + 13 * row, column=35).value = 'Mod Transition Count'
        sheet.cell(row=17 + 13 * row, column=35).value = 'Octant #'
        sheet.cell(row=16 + 13 * row, column=36).value = 'To'
        sheet.cell(row=18 + 13 * row, column=34).value = 'From'
        if q > mr:
            q = mr - 1
            sheet.cell(row=16 + 13 * row,
                       column=35).value = '{} - {}'.format(p, q + 1)
        else:
            sheet.cell(row=16 + 13 * row,
                       column=35).value = '{} - {}'.format(p, q)

        for col in range(8):
            sheet.cell(row=17 + 13 * row, column=36 + col).value = octant_key[col]
            sheet.cell(row=18 + 13 * row + col, column=35).value = octant_key[col]
            for k in range(8):
                sheet.cell(row=18 + 13 * row + k, column=36 + col).value = 0
        for p in range(p + 2, q + 2):
            index = 0
            count = 0
            val1 = sheet.cell(row=p, column=11).value
            val2 = sheet.cell(row=p + 1, column=11).value
            if val1 == '+1':
                index = 0
            elif val1 == '-1':
                index = 1
            elif val1 == '+2':
                index = 2
            elif val1 == '-2':
                index = 3
            elif val1 == '+3':
                index = 4
            elif val1 == '-3':
                index = 5
            elif val1 == '+4':
                index = 6
            elif val1 == '-4':
                index = 7
            if val2 == '+1':
                count = 0
            elif val2 == '-1':
                count = 1
            elif val2 == '+2':
                count = 2
            elif val2 == '-2':
                count = 3
            elif val2 == '+3':
                count = 4
            elif val2 == '-3':
                count = 5
            elif val2 == '+4':
                count = 6
            elif val2 == '-4':
                count = 7
            sheet.cell(row=18 + 13 * row + index, column=36 + count).value += 1
        p += mod
        q += mod
    borderSetFunc(sheet, 'AI3:AQ11')
    for row in range(iter):
        borderSetFunc(sheet, 'AI{}:AQ{}'.format(17 + row * 13, 17 + row * 13 + 8))
    k = 0
    for row in range(8):
        m = 0
        for col in range(8):
            val2 = sheet.cell(row=4 + row, column=36 + col).value
            if val2 > m:
                m = val2
                k = col
        sheet.cell(row=4 + row, column=36 + k).fill = yellow
    for p in range(iter):
        for row in range(8):
            m = 0
            for col in range(8):
                val2 = sheet.cell(row=18 + row + p * 13, column=36 + col).value
                if val2 > m:
                    m = val2
                    k = col
            sheet.cell(row=18 + row + p * 13, column=36 + k).fill = yellow


def octant_subsequence_range(sheet):
    mr = sheet.max_row
    sheet['AS1'] = 'Longest Subsequence Length'
    sheet['AS3'] = 'Octant ##'
    sheet['AT3'] = 'Longest Subsequence Length'
    sheet['AU3'] = 'Count'
    sub_seq_len = {'+1': 0, '-1': 0,           #dict to store sub seq lengths
           '+2': 0, '-2': 0,
           '+3': 0, '-3': 0,
           '+4': 0, '-4': 0}
    sub_seq_cnt = {'+1': 0, '-1': 0,                #dict to store sub seq cnt
             '+2': 0, '-2': 0,
             '+3': 0, '-3': 0,
             '+4': 0, '-4': 0}
    sub_seq_ind = {'+1': [0], '-1': [0],
                '+2': [0], '-2': [0],
                '+3': [0], '-3': [0],
                '+4': [0], '-4': [0]}
    t_range = {'+1': [], '-1': [],
                 '+2': [], '-2': [],
                 '+3': [], '-3': [],
                 '+4': [], '-4': []}
    temp = 0
    row = 2
    val_x = sheet.cell(row=2, column=11).value
    for row in range(2, mr + 1):
        val_x_i = sheet.cell(row=row, column=11).value
        if val_x_i != val_x:
            if temp > sub_seq_len[val_x]:
                sub_seq_len[val_x] = temp
                sub_seq_cnt[val_x] = 1
                sub_seq_ind[val_x].clear()
                sub_seq_ind[val_x].append(row - 1)
            elif temp == sub_seq_len[val_x]:
                sub_seq_cnt[val_x] += 1
                sub_seq_ind[val_x].append(row - 1)
            temp = 1
            val_x = val_x_i
        else:
            temp += 1
        row += 1
    for row in range(8):
        sheet.cell(row=4 + row, column=45).value = octant_key[row]
        sheet.cell(row=4 + row, column=46).value = sub_seq_len[octant_key[row]]
        sheet.cell(row=4 + row, column=47).value = sub_seq_cnt[octant_key[row]]
    for row in range(8):
        for j in sub_seq_ind[octant_key[row]]:
            index = j
            length = sub_seq_len[octant_key[row]]
            ini_val = sheet.cell(row=1 + index - length, column=1).value
            fin_val = sheet.cell(row=index, column=1).value
            t_range[octant_key[row]].append([ini_val, fin_val])
    sheet['AW1'] = 'Longest Subsequence Length With Range'
    sheet['AW3'] = "Octant ###"
    sheet['AX3'] = "Longest Subsequence Length"
    sheet['AY3'] = "Count"
    k = 4
    for row in octant_key:
        sheet.cell(row=k, column=49).value = row
        sheet.cell(row=k, column=50).value = sub_seq_len[row]
        sheet.cell(row=k, column=51).value = sub_seq_cnt[row]
        k = k + 1
        sheet.cell(row=k, column=49).value = "Time"
        sheet.cell(row=k, column=50).value = "From"
        sheet.cell(row=k, column=51).value = "To"
        k = k + 1
        for j in t_range[row]:
            sheet.cell(row=k, column=50).value = j[0]
            sheet.cell(row=k, column=51).value = j[1]
            k = k + 1
    borderSetFunc(sheet, 'AS3:AU11')
    borderSetFunc(sheet, 'AW3:AY{}'.format(k - 1))


def octant_analysis(mod=5000):
    for file in os.listdir():
        input_workbook = file
        output_workbook = generate_copy(input_workbook, mod)
        wb = openpyxl.load_workbook(output_workbook)
        active_sheet = wb.active
        octant_tagging(active_sheet)
        rank_and_count_func(mod, active_sheet)
        octant_transition_count(active_sheet)
        octant_subsequence_range(active_sheet)
        wb.save(output_workbook)


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod = 5000
octant_analysis(mod)


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
