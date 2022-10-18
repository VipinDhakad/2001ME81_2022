#Vipin Kumar Dhakad 2001ME81
from datetime import datetime
start_time = datetime.now()
import pandas as pd
import openpyxl
from datetime import datetime
start_time=datetime.now()
try:
    inp = openpyxl.load_workbook('octant_input.xlsx')
except:
    print("File not found!")
sheet_input = inp.active
##opening a new workbook to save as a output
output = openpyxl.Workbook()
out = output.active

##getting the size of the dataframe in terms of max rows and max columns
mr = sheet_input.max_row
mc = sheet_input.max_column

##copying the data of the input file into the newly created empty output file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# reading cell value from source input file
		data = sheet_input.cell(row = i, column = j)

		# writing the read value to destination output file
		out.cell(row = i, column = j).value = data.value

output.save("octant_output_ranking_excel.xlsx")
 
def octant_range_names(mod=5000):
    try:
        data=pd.read_excel('octant_output_ranking_excel.xlsx')
    except:
        print("File not found!")
    ##finding the mean of the column "U", "V", and "W" with the help of mean functions
    u_avg=data['U'].mean()
    v_avg=data['V'].mean()
    w_avg=data['W'].mean()
    ##writing the data to the first cells in the output file
    data.at[0,'U_avg']=u_avg
    data.at[0,'V_avg']=v_avg
    data.at[0,'W_avg']=w_avg
    i=0
    cnt_p1,cnt_n1,cnt_p2,cnt_n2,cnt_p3,cnt_n3,cnt_p4,cnt_n4=0,0,0,0,0,0,0,0
    ##calculating the values of deviation from the mean of every column
    for ele in data['U']:
        x=data.at[i,"U'=U-Uavg"]=data.at[i,'U']-u_avg
        y=data.at[i,"V'=V-Vavg"]=data.at[i,'V']-v_avg
        z=data.at[i,"W'=W-Wavg"]=data.at[i,'W']-w_avg
        ##tagging the octant begins here
        if x>0:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=1
                    cnt_p1=cnt_p1+1
                else:
                    data.at[i,'Octant']=-1
                    cnt_n1=cnt_n1+1
            else:
                if z>0:
                    data.at[i,'Octant']=4
                    cnt_p4=cnt_p4+1
                else:
                    data.at[i,'Octant']=-4
                    cnt_n4=cnt_n4+1
        else:
            if y>0:
                if z>0:
                    data.at[i,'Octant']=2
                    cnt_p2=cnt_p2+1
                else:
                    data.at[i,'Octant']=-2
                    cnt_n2=cnt_n2+1
            else:
                if z>0:
                    data.at[i,'Octant']=3
                    cnt_p3=cnt_p3+1
                else:
                    data.at[i,'Octant']=-3
                    cnt_n3=cnt_n3+1
        i=i+1
    data.at[0,'Octant ID']='Overall Count'   #writing the octant values obtained
    data.at[0,'1']=cnt_p1
    data.at[0,'-1']=cnt_n1
    data.at[0,'2']=cnt_p2
    data.at[0,'-2']=cnt_n2
    data.at[0,'3']=cnt_p3
    data.at[0,'-3']=cnt_n3
    data.at[0,'4']=cnt_p4
    data.at[0,'-4']=cnt_n4
  ##octant tagging for mod based range
    i=0
    prev=0
    iter=1
    cnt_p1,cnt_n1,cnt_p2,cnt_n2,cnt_p3,cnt_n3,cnt_p4,cnt_n4=0,0,0,0,0,0,0,0
    data.at[1,'Octant ID']=mod
    # print(len(data))
    while i<len(data):
        flag=False
        for j in range(prev,mod*iter):
            if i<len(data):
                if data.at[i,'Octant']==1:
                    cnt_p1=cnt_p1+1
                if data.at[i,'Octant']==-1:
                    cnt_n1=cnt_n1+1
                if data.at[i,'Octant']==2:
                    cnt_p2=cnt_p2+1
                if data.at[i,'Octant']==-2:
                    cnt_n2=cnt_n2+1
                if data.at[i,'Octant']==3:
                    cnt_p3=cnt_p3+1
                if data.at[i,'Octant']==-3:
                    cnt_n3=cnt_n3+1
                if data.at[i,'Octant']==4:
                    cnt_p4=cnt_p4+1
                if data.at[i,'Octant']==-4:
                    cnt_n4=cnt_n4+1
                i=i+1
            else:
                flag=True
        data.at[iter+1,'Octant ID']=str(prev)+"-"+str(mod*iter-1)
        if flag:
            data.at[iter+1,'Octant ID']=str(prev)+"-"+str(len(data))
        data.at[iter+1,'1']=cnt_p1         ##writing the values obtained
        data.at[iter+1,'-1']=cnt_n1
        data.at[iter+1,'2']=cnt_p2
        data.at[iter+1,'-2']=cnt_n2
        data.at[iter+1,'3']=cnt_p3
        data.at[iter+1,'-3']=cnt_n3
        data.at[iter+1,'4']=cnt_p4
        data.at[iter+1,'-4']=cnt_n4
        cnt_p1,cnt_n1,cnt_p2,cnt_n2,cnt_p3,cnt_n3,cnt_p4,cnt_n4=0,0,0,0,0,0,0,0
        prev=mod*iter
        iter=iter+1

    print("Data writing complete")
    ##saving the output file generated
    data.to_excel('octant_output_ranking_excel.xlsx',index=False)
    print("Starting rank calculation")
    data=pd.read_excel('octant_output_ranking_excel.xlsx')
    octants=[1,-1,2,-2,3,-3,4,-4]
    i=1
    list_of_octant_counts=[]
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    for octant in octants:
        data.at[0,f'Rank of {octant}']=f'Rank {i}'
        i+=1
        list_of_octant_counts.append((data.at[0,f'{octant}'],octant))   #creating a list with count and octant id
    list_of_octant_counts.sort(reverse=True)          ##sorting the list in descending order
    i=1
    for item in list_of_octant_counts:
        data.at[1,f'Rank of {item[1]}']=i
        i+=1
    data.at[1,'Rank1 Octant ID']=list_of_octant_counts[0][1]
    data.at[1,'Rank 1 Octant Name']=octant_name_id_mapping[f"{list_of_octant_counts[0][1]}"]         ##setting the octant name
    row=2
    Rank1_mod_values_count=[]   ##finding the rank1 octant for mod based range values
    while row-1<iter:
        list_of_octant_counts.clear()
        for octant in octants:
            list_of_octant_counts.append((data.at[row,f'{octant}'],octant))
        list_of_octant_counts.sort(reverse=True)
        i=1
        for item in list_of_octant_counts:
            data.at[row,f'Rank of {item[1]}']=i
            i+=1
        data.at[row,'Rank1 Octant ID']=list_of_octant_counts[0][1]
        data.at[row,'Rank 1 Octant Name']=octant_name_id_mapping[f"{list_of_octant_counts[0][1]}"]
        Rank1_mod_values_count.append(list_of_octant_counts[0][1])
        row+=1
    print('Finished Calculations')
    row=13                          #generating the summary of rank1 values
    data.iloc[12+iter,12]='Octant ID'
    data.iloc[12+iter,13]='Octant Name'
    data.iloc[12+iter,14]='Count of Rank 1 MOD Values'
    for octant in octants:
        data.at[row+iter,'1']=octant
        data.at[row+iter,'-1']=octant_name_id_mapping[f"{octant}"]
        data.at[row+iter,'2']=Rank1_mod_values_count.count(octant)
        row+=1
    data.to_excel('octant_output_ranking_excel.xlsx',index=False)        #saving the output


from platform import python_version
ver = python_version()

if ver == "3.8.10":              #checking the version of python installed
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=3000
octant_range_names(mod)



#This shall be the last lines of the code ... and these are
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
