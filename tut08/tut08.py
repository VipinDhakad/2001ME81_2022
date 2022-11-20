# Vipin Kumar Dhakad
# 2001ME81
from platform import python_version
import os
import openpyxl 
import pandas as pd 
from datetime import datetime

 
start_time = datetime.now() 

india_innings = open("india_inns2.txt", "r+") 
pakistan_innings = open("pak_inns1.txt", "r+") 
teams = open("teams.txt", "r+") 

team_input = teams.readlines() 
team_pakistan = team_input[0] 
player_pakistan = team_pakistan[23:-1:].split(",")

 
team_india = team_input[2] 
player_india = team_india[20:-1:].split(",") 

line_india = india_innings.readlines()

 
for l_index in line_india: 
    if l_index == '\n': 
        line_india.remove(l_index) 

line_pakistan = pakistan_innings.readlines()

 
for l_index in line_pakistan: 
    if l_index == '\n': 
        line_pakistan.remove(l_index) 

wb = openpyxl.Workbook() 
sheet = wb.active 

Wic_ind, Wic_Pak, India_b, Pakistan_b, tB_ind, tB_Pak = 0, 0, 0, 0, 0, 0 

out_Ind = {} 
out_Pak = {} 
iB = {} 
pB = {} 
bat_ind = {} 
bat_pak = {} 

for l_index in line_pakistan: 
    index_line = l_index.index(".") 
    pakistan_overs = l_index[0:index_line + 2] 
    var = l_index[index_line + 2::].split(",") 
    bowler = var[0].split("to") 
    if f"{bowler[0].strip()}" not in iB.keys(): 
        iB[f"{bowler[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0] 

    elif "wide" in var[1]: 
        pass 

    elif "bye" in var[1]: 
        if "FOUR" in var[2]: 
            Pakistan_b += 4 
        elif "1" in var[2]: 
            Pakistan_b += 1 
        elif "2" in var[2]: 
            Pakistan_b += 2 
        elif "3" in var[2]: 
            Pakistan_b += 3 
        elif "4" in var[2]: 
            Pakistan_b += 4 
        elif "5" in var[2]: 
            Pakistan_b += 5 

    else: 
        iB[f"{bowler[0].strip()}"][0] += 1 

    if f"{bowler[1].strip()}" not in bat_pak.keys() and var[1] != "wide": 
        bat_pak[f"{bowler[1].strip()}"] = [0, 1, 0, 0, 0] 

    elif "wide" in var[1]: 
        pass 

    else: 
        bat_pak[f"{bowler[1].strip()}"][1] += 1 

    if "out" in var[1]: 
        iB[f"{bowler[0].strip()}"][3] += 1 
        if "Bowled" in var[1].split("!!")[0]: 
            out_Pak[f"{bowler[1].strip()}"] = ("b" + bowler[0]) 
        elif "Caught" in var[1].split("!!")[0]: 
            wicket_by = (var[1].split("!!")[0]).split("by") 
            out_Pak[f"{bowler[1].strip()}"] = ( 
                "c" + wicket_by[1] + " b " + bowler[0]) 
        elif "Lbw" in var[1].split("!!")[0]: 
            out_Pak[f"{bowler[1].strip()}"] = ("lbw  b "+bowler[0]) 

    if "no run" in var[1] or "out" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 0 
        bat_pak[f"{bowler[1].strip()}"][0] += 0 

    elif "1 run" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 1 
        bat_pak[f"{bowler[1].strip()}"][0] += 1 

    elif "2 run" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 2 
        bat_pak[f"{bowler[1].strip()}"][0] += 2 

    elif "3 run" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 3 
        bat_pak[f"{bowler[1].strip()}"][0] += 3 

    elif "4 run" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 4 
        bat_pak[f"{bowler[1].strip()}"][0] += 4 

    elif "FOUR" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 4 
        bat_pak[f"{bowler[1].strip()}"][0] += 4 
        bat_pak[f"{bowler[1].strip()}"][2] += 1 

    elif "SIX" in var[1]: 
        iB[f"{bowler[0].strip()}"][2] += 6 
        bat_pak[f"{bowler[1].strip()}"][0] += 6 
        bat_pak[f"{bowler[1].strip()}"][3] += 1 

    elif "wide" in var[1]: 
        if "wides" in var[1]: 
            iB[f"{bowler[0].strip()}"][2] += int(var[1][1]) 
            iB[f"{bowler[0].strip()}"][5] += int(var[1][1]) 
        else: 
            iB[f"{bowler[0].strip()}"][2] += 1 
            iB[f"{bowler[0].strip()}"][5] += 1

 
for l_index in bat_pak.values(): 
    l_index[-1] = round((l_index[0]/l_index[1])*100, 2) 

    
for l_index in line_india: 
    index_line = l_index.index(".") 
    indOver = l_index[0:index_line+2] 
    var = l_index[index_line + 2::].split(",") 
    bowler = var[0].split("to") 
    if f"{bowler[0].strip()}" not in pB.keys(): 
        pB[f"{bowler[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0] 

    elif "wide" in var[1]: 
        pass 

    elif "bye" in var[1]: 
        if "FOUR" in var[2]: 
            India_b += 4 
        elif "1" in var[2]: 
            India_b += 1 
        elif "2" in var[2]: 
            India_b += 2 
        elif "3" in var[2]: 
            India_b += 3 
        elif "4" in var[2]: 
            India_b += 4 
        elif "5" in var[2]: 
            India_b += 5 

    else: 
        pB[f"{bowler[0].strip()}"][0] += 1 

    if f"{bowler[1].strip()}" not in bat_ind.keys() and var[1] != "wide": 
        bat_ind[f"{bowler[1].strip()}"] = [0, 1, 0, 0, 0] 

    elif "wide" in var[1]: 
        pass 

    else: 
        bat_ind[f"{bowler[1].strip()}"][1] += 1 

    if "out" in var[1]: 
        pB[f"{bowler[0].strip()}"][3] += 1 
        if "Bowled" in var[1].split("!!")[0]: 
            out_Ind[f"{bowler[1].strip()}"] = ("b" + bowler[0]) 
        elif "Caught" in var[1].split("!!")[0]: 
            wicket_by = (var[1].split("!!")[0]).split("by") 
            out_Ind[f"{bowler[1].strip()}"] = ( 
                "c" + wicket_by[1] + " b " + bowler[0]) 
        elif "Lbw" in var[1].split("!!")[0]: 
            out_Ind[f"{bowler[1].strip()}"] = ("lbw  b " + bowler[0]) 

    if "no run" in var[1] or "out" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 0 
        bat_ind[f"{bowler[1].strip()}"][0] += 0 

    elif "1 run" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 1 
        bat_ind[f"{bowler[1].strip()}"][0] += 1 

    elif "2 run" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 2 
        bat_ind[f"{bowler[1].strip()}"][0] += 2 

    elif "3 run" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 3 
        bat_ind[f"{bowler[1].strip()}"][0] += 3 

    elif "4 run" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 4 
        bat_ind[f"{bowler[1].strip()}"][0] += 4 

    elif "FOUR" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 4 
        bat_ind[f"{bowler[1].strip()}"][0] += 4 
        bat_ind[f"{bowler[1].strip()}"][2] += 1 

    elif "SIX" in var[1]: 
        pB[f"{bowler[0].strip()}"][2] += 6 
        bat_ind[f"{bowler[1].strip()}"][0] += 6 
        bat_ind[f"{bowler[1].strip()}"][3] += 1 

    elif "wide" in var[1]: 
        if "wides" in var[1]: 
            pB[f"{bowler[0].strip()}"][2] += int(var[1][1]) 
            pB[f"{bowler[0].strip()}"][5] += int(var[1][1]) 
        else: 
            pB[f"{bowler[0].strip()}"][2] += 1 
            pB[f"{bowler[0].strip()}"][5] += 1 

for batter_index in bat_ind.values(): 
    batter_index[-1] = round((batter_index[0]/batter_index[1])*100, 2) 

for batter_index in bat_pak.values(): 
    batter_index[-1] = round((batter_index[0]/batter_index[1])*100, 2) 

for batter_index in iB.values(): 
    if batter_index[0] % 6 == 0: 
        batter_index[0] = batter_index[0]//6 
    else: 
        batter_index[0] = (batter_index[0]//6) + (batter_index[0] % 6)/10 

for batter_index in pB.values(): 
    if batter_index[0] % 6 == 0: 
        batter_index[0] = batter_index[0]//6 
    else: 
        batter_index[0] = (batter_index[0]//6) + (batter_index[0] % 6)/10 

for batter_index in iB.values(): 
    index_line = str(batter_index[0]) 
    if "." in index_line: 
        balls = int(index_line[0])*6 + int(index_line[2]) 
        batter_index[-1] = round((batter_index[2]/balls)*6, 1) 
    else: 
        batter_index[-1] = round((batter_index[2]/batter_index[0]), 1) 

for batter_index in pB.values(): 
    index_line = str(batter_index[0]) 
    if "." in index_line: 
        balls = int(index_line[0])*6 + int(index_line[2]) 
        batter_index[-1] = round((batter_index[2]/balls)*6, 1) 
    else: 
        batter_index[-1] = round((batter_index[2]/batter_index[0]), 1) 

pakBatters = [] 

for name in bat_pak.keys(): 
    pakBatters.append(name) 

for l_index in range(len(bat_pak)): 
    sheet.cell(5 + l_index, 1).value = pakBatters[l_index] 
    sheet.cell(5 + l_index, 5).value = bat_pak[pakBatters[l_index]][0] 
    sheet.cell(5 + l_index, 6).value = bat_pak[pakBatters[l_index]][1] 
    sheet.cell(5 + l_index, 7).value = bat_pak[pakBatters[l_index]][2] 
    sheet.cell(5 + l_index, 8).value = bat_pak[pakBatters[l_index]][3] 
    sheet.cell(5 + l_index, 9).value = bat_pak[pakBatters[l_index]][4] 
    if pakBatters[l_index] not in out_Pak: 
        sheet.cell(5+l_index, 3).value = "not out" 
    else: 
        sheet.cell(5+l_index, 3).value = out_Pak[pakBatters[l_index]] 


sheet.cell(3, 1).value = "BATTERS" 
sheet["E3"] = "RUNS" 
sheet["F3"] = "BALLS" 
sheet["G3"] = " 4s " 
sheet["H3"] = " 6s " 
sheet["I3"] = "  SR  " 
sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER" 
sheet["D18"] = "MAIDEN" 
sheet["E18"] = "RUNS" 
sheet["F18"] = "WICKET" 
sheet["G18"] = "NO-BALL" 
sheet["H18"] = "WIDE" 
sheet["I18"] = "ECONOMY" 


pakBowlers = [] 

for name in pB.keys(): 
    pakBowlers.append(name) 

for l_index in range(len(pB)): 
    sheet.cell(42 + l_index, 1).value = pakBowlers[l_index] 
    sheet.cell(42 + l_index, 3).value = pB[pakBowlers[l_index]][0] 
    sheet.cell(42 + l_index, 4).value = pB[pakBowlers[l_index]][1] 
    sheet.cell(42 + l_index, 5).value = pB[pakBowlers[l_index]][2] 
    sheet.cell(42 + l_index, 6).value = pB[pakBowlers[l_index]][3] 
    sheet.cell(42 + l_index, 7).value = pB[pakBowlers[l_index]][4] 
    sheet.cell(42 + l_index, 8).value = pB[pakBowlers[l_index]][5] 
    sheet.cell(42 + l_index, 9).value = pB[pakBowlers[l_index]][6] 
    tB_Pak += pB[pakBowlers[l_index]][2] 
    Wic_ind += pB[pakBowlers[l_index]][3] 


sheet.cell(11 + len(bat_pak) + len(pB), 1).value = "# INDIA" 
sheet.cell(11 + len(bat_pak) + len(pB), 2).value = " INNINGS" 
indBatters = [] 

for name in bat_ind.keys(): 
    indBatters.append(name) 

for l_index in range(len(bat_ind)): 
    sheet.cell(31 + l_index, 1).value = indBatters[l_index] 
    sheet.cell(31 + l_index, 5).value = bat_ind[indBatters[l_index]][0] 
    sheet.cell(31 + l_index, 6).value = bat_ind[indBatters[l_index]][1] 
    sheet.cell(31 + l_index, 7).value = bat_ind[indBatters[l_index]][2] 
    sheet.cell(31 + l_index, 8).value = bat_ind[indBatters[l_index]][3] 
    sheet.cell(31 + l_index, 9).value = bat_ind[indBatters[l_index]][4] 
    if indBatters[l_index] not in out_Ind: 
        sheet.cell(31 + l_index, 3).value = "not out" 

    else: 
        sheet.cell(31 + l_index, 3).value = out_Ind[indBatters[l_index]] 


sheet["A29"] = "BATTERS" 
sheet["E29"] = "RUNS" 
sheet["F29"] = "BALLS" 
sheet["G29"] = " 4s " 
sheet["H29"] = " 6s " 
sheet["I29"] = "  SR  " 
sheet["A40"] = "BOWLER" 
sheet["C40"] = "OVER" 
sheet["D40"] = "MAIDEN" 
sheet["E40"] = "RUNS" 
sheet["F40"] = "WICKET" 
sheet["G40"] = "NO-BALL" 
sheet["H40"] = "WIDE" 
sheet["I40"] = "ECONOMY" 
indBowlers = [] 

for name in iB.keys(): 
    indBowlers.append(name) 


for l_index in range(len(iB)): 
    sheet.cell(20 + l_index, 1).value = indBowlers[l_index] 
    sheet.cell(20 + l_index, 3).value = iB[indBowlers[l_index]][0] 
    sheet.cell(20 + l_index, 4).value = iB[indBowlers[l_index]][1] 
    sheet.cell(20 + l_index, 5).value = iB[indBowlers[l_index]][2] 
    sheet.cell(20 + l_index, 6).value = iB[indBowlers[l_index]][3] 
    sheet.cell(20 + l_index, 7).value = iB[indBowlers[l_index]][4] 
    sheet.cell(20 + l_index, 8).value = iB[indBowlers[l_index]][5] 
    sheet.cell(20 + l_index, 9).value = iB[indBowlers[l_index]][6] 
    tB_ind += iB[indBowlers[l_index]][2] 
    Wic_Pak += iB[indBowlers[l_index]][3] 


# Total Score 
indiaS = tB_ind + Pakistan_b 
pakistanS = tB_Pak + India_b 

sheet["E27"] = " " + str(indiaS) + " - " + str(Wic_ind) 
sheet["F27"] = str(indOver) 

axis1 = " " + str(pakistanS) + " - " + str(Wic_Pak) 
axis2 = str(pakistan_overs) 
# Writing the output in csv format 

wb.save("Scoreboard.xlsx") 

df = pd.read_excel('Scoreboard.xlsx') 
df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" ", " ", axis1, axis2, " ", " ", " "], axis='columns') 
df.to_csv('Scorecard.csv', index=False) 

try: 
    os.path.exists("Scoreboard.xlsx") 
    os.remove("Scoreboard.xlsx") 

except: 
    print("Unexpected Error!") 

ver = python_version() 

if ver == "3.8.10": 
    print("Correct Version Installed") 

else: 
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw") 

end_time = datetime.now() 
print('Duration of Program Execution: {}'.format(end_time - start_time))
